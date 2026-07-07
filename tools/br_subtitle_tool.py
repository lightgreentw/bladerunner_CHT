#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Blade Runner 字幕 & UI 翻譯工具 v2
================================
指令：
  dump       — 掃描遊戲目錄，輸出 subtitles.xlsx（字幕）+ ui_text.xlsx（介面）
  import     — 讀入兩個 Excel，自動偵測 SUBFONT.TTF / UIFONT.TTF，輸出 SUBTITLES.MIX + UI .TRE
  patch-tlk  — 直接把 UI 翻譯寫回 TLK 檔案（原地修改，自動備份）

使用範例：
  python3 br_subtitle_tool.py dump   --dir "C:\\game" --out .
  python3 br_subtitle_tool.py import --subxlsx subtitles.xlsx --uixlsx ui_text.xlsx --out output/
  python3 br_subtitle_tool.py patch-tlk --dir "C:\\game" --xlsx ui_text.xlsx
"""

import struct, os, sys, argparse, shutil
from datetime import datetime

# ─── MIX ──────────────────────────────────────────────────────────────────────


def fold_hash(filename: str) -> int:
    s = filename.upper()
    i = 0
    h = 0
    while i < len(s) and i < 12:
        g = 0
        for _ in range(4):
            g >>= 8
            if i < len(s):
                g |= ord(s[i]) << 24
                i += 1
        h = ((h << 1) | ((h >> 31) & 1)) + g
        h &= 0xFFFFFFFF
    return h


def hash_signed(h):
    return struct.unpack("<i", struct.pack("<I", h))[0]


def unpack_mix(data: bytes) -> dict:
    """→ {uint32_hash: bytes}"""
    n = struct.unpack_from("<H", data, 0)[0]
    ds = 2 + 4 + n * 12
    out = {}
    for i in range(n):
        b = 6 + i * 12
        h = struct.unpack_from("<I", data, b)[0]
        off = struct.unpack_from("<I", data, b + 4)[0]
        sz = struct.unpack_from("<I", data, b + 8)[0]
        out[h] = data[ds + off : ds + off + sz]
    return out


def pack_mix(file_map: dict) -> bytes:
    """file_map = {name_str: bytes} — entry table 按 int32 hash 排序"""
    recs = []
    ds_off = 0
    for name, d in file_map.items():
        u = fold_hash(name)
        recs.append((u, hash_signed(u), ds_off, len(d), d))
        ds_off += len(d)
    sorted_recs = sorted(recs, key=lambda r: r[1])
    buf = struct.pack("<H", len(recs)) + struct.pack("<I", ds_off)
    for u, s, off, sz, _ in sorted_recs:
        buf += struct.pack("<III", u, off, sz)
    for _, _, _, _, d in recs:
        buf += d
    return buf


def repack_mix_preserve_unknown(original_data: bytes, updated_entries: dict) -> bytes:
    """重新封裝 MIX，保留所有原始 entry（含未知 hash），只替換有更新的部分"""
    num_entries = struct.unpack_from("<H", original_data, 0)[0]
    data_start = 2 + 4 + num_entries * 12

    original_items = []
    for i in range(num_entries):
        base = 6 + i * 12
        h = struct.unpack_from("<I", original_data, base)[0]
        off = struct.unpack_from("<I", original_data, base + 4)[0]
        sz = struct.unpack_from("<I", original_data, base + 8)[0]
        if h in updated_entries:
            entry_data = updated_entries[h]
        else:
            entry_data = original_data[data_start + off : data_start + off + sz]
        original_items.append((h, entry_data))

    records = []
    ds_off = 0
    for h, data in original_items:
        records.append((h, hash_signed(h), ds_off, len(data), data))
        ds_off += len(data)

    sorted_records = sorted(records, key=lambda r: r[1])
    buf = struct.pack("<H", len(records))
    buf += struct.pack("<I", ds_off)
    for h, s, off, sz, _ in sorted_records:
        buf += struct.pack("<III", h, off, sz)
    for _, _, _, _, data in records:
        buf += data
    return buf


# ─── TRE ──────────────────────────────────────────────────────────────────────


def read_tre(data: bytes, encoding="utf-8") -> list:
    """→ [(id, text), ...]"""
    count = struct.unpack_from("<I", data, 0)[0]
    ids = [struct.unpack_from("<I", data, 4 + i * 4)[0] for i in range(count)]
    offs = [
        struct.unpack_from("<I", data, 4 + count * 4 + i * 4)[0]
        for i in range(count + 1)
    ]
    pos = 4 + count * 4 + (count + 1) * 4
    ss = pos - 4
    blob = data[pos:]
    result = []
    for i in range(count):
        raw = blob[offs[i] - ss : offs[i + 1] - ss]
        result.append((ids[i], raw.rstrip(b"\x00").decode(encoding, errors="replace")))
    return result


def write_tre(entries: list, encoding="utf-8") -> bytes:
    """entries = [(id, text), ...]"""
    count = len(entries)
    encoded = [t.encode(encoding) + b"\x00" for _, t in entries]
    base = count * 4 + (count + 1) * 4
    offs = []
    pos = 0
    for enc in encoded:
        offs.append(base + pos)
        pos += len(enc)
    offs.append(base + pos)
    buf = struct.pack("<I", count)
    for eid, _ in entries:
        buf += struct.pack("<I", eid)
    for o in offs:
        buf += struct.pack("<I", o)
    for enc in encoded:
        buf += enc
    return buf


# ─── 字幕 TRE 定義（SUBTITLES.MIX 內容） ───────────────────────────────────────

LANG = "E"  # 英文版

TRE_DEFS = [
    ("INGQUO", "ingquo", True),
    ("WSTLGO", "vqa", False),
    ("BRLOGO", "vqa", False),
    ("INTRO", "vqa", True),
    ("MW_A", "vqa", True),
    ("MW_B01", "vqa", True),
    ("MW_B02", "vqa", True),
    ("MW_B03", "vqa", True),
    ("MW_B04", "vqa", True),
    ("MW_B05", "vqa", True),
    ("INTRGT", "vqa", True),
    ("MW_C01", "vqa", True),
    ("MW_C02", "vqa", True),
    ("MW_C03", "vqa", True),
    ("MW_D", "vqa", True),
    ("END04A", "vqa", True),
    ("END04B", "vqa", True),
    ("END04C", "vqa", True),
    ("END06", "vqa", True),
    ("END01A", "vqa", True),
    ("END01B", "vqa", True),
    ("END01C", "vqa", True),
    ("END01D", "vqa", True),
    ("END01E", "vqa", True),
    ("END01F", "vqa", True),
    ("END03", "vqa", True),
    ("TB_FLY", "vqa", True),
    ("SBTLVERS", "meta", False),
    ("EXTRA", "meta", False),
    ("SUBTLS_E.FON", "binary", False),
]


def tre_filename(prefix, localized, lang=LANG):
    if not localized:
        return (
            f"{prefix}_E.TRE"
            if prefix not in ("SBTLVERS", "EXTRA", "SUBTLS_E.FON")
            else (prefix if "." in prefix else f"{prefix}.TRE")
        )
    return f"{prefix}_{lang}.TR{lang}"


ALL_KNOWN_SUB = []
for prefix, typ, loc in TRE_DEFS:
    if typ == "binary":
        ALL_KNOWN_SUB.append(prefix)
    else:
        ALL_KNOWN_SUB.append(tre_filename(prefix, loc))

HASH_TO_NAME_SUB = {fold_hash(f): f for f in ALL_KNOWN_SUB}


def unpack_vqa_id(packed_id):
    frame_start = packed_id & 0x0000FFFF
    frame_end = (packed_id >> 16) & 0x0000FFFF
    return frame_start, frame_end


def pack_vqa_id(frame_start, frame_end):
    return (frame_end << 16) | (frame_start & 0xFFFF)


# ─── UI TRE 定義（存於 1.TLK / 2.TLK / 3.TLK / A.TLK） ─────────────────────────

UI_TRES = [
    ("KIA", "KIA 介面（按鈕 tooltip、章節標題、分類標籤）"),
    ("HELP", "KIA 說明頁面（Gameplay/Combat/Keyboard 說明文字）"),
    ("OPTIONS", "設定選單（音樂、音效、字幕、語言等）"),
    ("SPINDEST", "飛行車目的地名稱"),
    ("VK", "Voight-Kampff介面"),
    ("ACTORS", "角色名稱"),
    ("CRIMES", "案件名稱"),
    ("CLUETYPE", "線索類型標籤"),
    ("CLUES", '線索名稱（KIA 右側線索清單，如 "Maggie" Bracelet）'),
    ("DLGMENU", "遊戲中對話選項名稱（如 VOIGT-KAMPFF / DONE / CRYSTAL 等）"),
]

TLK_NAMES = ["1.TLK", "2.TLK", "3.TLK", "A.TLK", "STARTUP.MIX"]

# 字型固定檔名（對應 bladerunner.cpp 的修改：UIFONT.TTF 為介面專用字型）
SUB_FONT_FILENAME = (
    "SUBFONT.TTF"  # 字幕專用字型（寫入 SUBTITLES.MIX 的 SBTLVERS.TRE fontName）
)
UI_FONT_FILENAME = (
    "UIFONT.TTF"  # 介面專用字型（固定檔名 UIFONT.TTF，寫入 SUBTITLES.MIX）
)

# ─── Excel 共用樣式 helper ──────────────────────────────────────────────────────


def _get_styles():
    from openpyxl.styles import Font as XLFont, PatternFill, Alignment, Border, Side

    return {
        "hdr_font": XLFont(bold=True, color="FFFFFF", size=11),
        "hdr_fill": PatternFill("solid", fgColor="1F3864"),
        "wrap": Alignment(wrap_text=True, vertical="top"),
        "thin": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


# ─── DUMP：統一指令，掃描整個遊戲目錄 ───────────────────────────────────────────


def cmd_dump(game_dir: str, out_dir: str):
    """
    掃描 game_dir 下的 SUBTITLES.MIX 與 1.TLK/2.TLK/3.TLK/A.TLK，
    輸出兩個 Excel：
      {out_dir}/subtitles.xlsx — 字幕（INGQUO + 所有 VQA 過場字幕）
      {out_dir}/ui_text.xlsx   — 介面文字（KIA/OPTIONS/SPINDEST/...）
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ 請先安裝：pip install openpyxl")
        sys.exit(1)

    os.makedirs(out_dir, exist_ok=True)
    styles = _get_styles()

    # ── 1. 找 SUBTITLES.MIX ──────────────────────────────────────────────
    mix_path = os.path.join(game_dir, "SUBTITLES.MIX")
    has_subtitles = os.path.isfile(mix_path)

    # ── 2. 找 TLK 檔案 ───────────────────────────────────────────────────
    tlk_files = [
        os.path.join(game_dir, n)
        for n in TLK_NAMES
        if os.path.isfile(os.path.join(game_dir, n))
    ]

    if not has_subtitles and not tlk_files:
        print(f"❌ 在 {game_dir} 找不到 SUBTITLES.MIX 或任何 TLK 檔案")
        sys.exit(1)

    print(f"[1/4] 掃描遊戲目錄：{game_dir}")
    print(
        f"      SUBTITLES.MIX：{'找到 ✅' if has_subtitles else '未找到（將跳過字幕匯出）'}"
    )
    print(
        f"      TLK 檔案：{[os.path.basename(f) for f in tlk_files] if tlk_files else '未找到（將跳過介面匯出）'}"
    )

    sub_xlsx_path = os.path.join(out_dir, "subtitles.xlsx")
    ui_xlsx_path = os.path.join(out_dir, "ui_text.xlsx")

    # ── 3. 匯出字幕 subtitles.xlsx ──────────────────────────────────────
    if has_subtitles:
        print(f"\n[2/4] 匯出字幕 → {sub_xlsx_path}")
        with open(mix_path, "rb") as f:
            mix_data = f.read()
        entries = unpack_mix(mix_data)
        _export_subtitles_xlsx(
            entries, sub_xlsx_path, styles, get_column_letter, openpyxl
        )
    else:
        print(f"\n[2/4] 跳過字幕匯出（找不到 SUBTITLES.MIX）")

    # ── 4. 匯出介面 ui_text.xlsx ─────────────────────────────────────────
    if tlk_files:
        print(f"\n[3/4] 匯出介面文字 → {ui_xlsx_path}")
        all_entries = {}
        for tlk_path in tlk_files:
            with open(tlk_path, "rb") as f:
                data = f.read()
            all_entries.update(unpack_mix(data))
        _export_ui_xlsx(
            all_entries, game_dir, ui_xlsx_path, styles, get_column_letter, openpyxl
        )
    else:
        print(f"\n[3/4] 跳過介面匯出（找不到 TLK 檔案）")

    print(f"\n[4/4] 完成！")
    if has_subtitles:
        print(f"      {sub_xlsx_path}")
    if tlk_files:
        print(f"      {ui_xlsx_path}")


def _export_subtitles_xlsx(entries, out_path, styles, get_column_letter, openpyxl):
    from openpyxl.styles import PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ingquo_fill = PatternFill("solid", fgColor="E8F0FE")
    vqa_fill = PatternFill("solid", fgColor="FEF9E7")

    total_rows = 0
    sheet_stats = []

    for prefix, typ, loc in TRE_DEFS:
        if typ in ("binary", "meta"):
            continue

        fname = tre_filename(prefix, loc)
        h = fold_hash(fname)
        if h not in entries:
            alt = f"{prefix}.TRE"
            h2 = fold_hash(alt)
            if h2 not in entries:
                continue
            h = h2
            fname = alt

        rows = read_tre(entries[h])
        if not rows:
            continue

        ws = wb.create_sheet(title=prefix[:31])
        ws.sheet_view.showGridLines = True

        if typ == "ingquo":
            headers = ["Quote ID", "英文原文", "中文翻譯", "備註"]
            col_widths = [12, 60, 60, 20]
        else:
            headers = [
                "Frame Start",
                "Frame End",
                "Packed ID (hex)",
                "英文原文",
                "中文翻譯",
                "備註",
            ]
            col_widths = [12, 12, 18, 60, 60, 20]

        for ci, (h_txt, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h_txt)
            cell.font = styles["hdr_font"]
            cell.fill = styles["hdr_fill"]
            cell.alignment = styles["wrap"]
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        row_fill = ingquo_fill if typ == "ingquo" else vqa_fill
        meaningful = 0
        for ri, (eid, txt) in enumerate(rows, 2):
            if txt.strip():
                meaningful += 1
            if typ == "ingquo":
                ws.cell(row=ri, column=1, value=eid)
                ws.cell(row=ri, column=2, value=txt)
                ws.cell(row=ri, column=3, value="")
                ws.cell(row=ri, column=4, value="")
            else:
                fs, fe = unpack_vqa_id(eid)
                ws.cell(row=ri, column=1, value=fs)
                ws.cell(row=ri, column=2, value=fe)
                ws.cell(row=ri, column=3, value=f"0x{eid:08X}")
                ws.cell(row=ri, column=4, value=txt)
                ws.cell(row=ri, column=5, value="")
                ws.cell(row=ri, column=6, value="")

            for ci in range(1, len(headers) + 1):
                cell = ws.cell(row=ri, column=ci)
                cell.fill = row_fill
                cell.alignment = styles["wrap"]
                cell.border = styles["thin"]
                ws.row_dimensions[ri].height = 40

        total_rows += len(rows)
        sheet_stats.append((prefix, typ, len(rows), meaningful))
        print(f"      ✅ {fname:25s} {len(rows):4d} 條")

    # 說明頁
    from openpyxl.styles import Font as XLFont

    ws_info = wb.create_sheet(title="📖 使用說明", index=0)
    ws_info.column_dimensions["A"].width = 80
    info_lines = [
        ("Blade Runner 字幕翻譯工作表", XLFont(bold=True, size=14)),
        (f'產生時間：{datetime.now().strftime("%Y-%m-%d %H:%M")}', XLFont(size=11)),
        ("", None),
        (
            "【填寫】INGQUO 是遊戲對話字幕，其他是過場動畫字幕；中文翻譯欄留空則保留英文",
            XLFont(size=11),
        ),
        (
            "【匯入】python br_subtitle_tool.py import --subxlsx 此檔案 --uixlsx ui_text.xlsx --out 輸出目錄",
            XLFont(size=11),
        ),
        ("", None),
        ("【工作表統計】", XLFont(bold=True, size=12)),
    ]
    for ri, (text, font) in enumerate(info_lines, 1):
        cell = ws_info.cell(row=ri, column=1, value=text)
        if font:
            cell.font = font
    for prefix, typ, total, meaningful in sheet_stats:
        ri += 1
        ws_info.cell(
            row=ri,
            column=1,
            value=f"  {prefix:15s}  {total:4d} 條（有內容：{meaningful:4d}）",
        )

    wb.save(out_path)
    print(f"      總計：{total_rows} 條台詞")


def _export_ui_xlsx(
    all_entries, game_dir, out_path, styles, get_column_letter, openpyxl
):
    from openpyxl.styles import PatternFill, Font as XLFont

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    row_fill = PatternFill("solid", fgColor="FFF9E6")

    ws_info = wb.create_sheet("📖 說明", 0)
    ws_info.column_dimensions["A"].width = 80
    for ri, text in enumerate(
        [
            "Blade Runner 遊戲介面翻譯工作表",
            f"來源：{game_dir}",
            "",
            "【填寫】在「中文翻譯」欄填入繁體中文，空白處保留英文",
            "【匯入】python br_subtitle_tool.py import --subxlsx subtitles.xlsx --uixlsx 此檔案 --out 輸出目錄",
            "【生效方式 A】import 後把 .TRE 檔案複製到遊戲目錄（散落檔案覆蓋）",
            "【生效方式 B】用 patch-tlk 指令直接寫回 TLK（更乾淨，見 patch-tlk --help）",
        ],
        1,
    ):
        ws_info.cell(row=ri, column=1, value=text)

    total = 0
    for tre_name, description in UI_TRES:
        fname = f"{tre_name}.TRE"
        h = fold_hash(fname)
        if h not in all_entries:
            print(f"      ⚠️  找不到 {fname}")
            continue

        rows = read_tre(all_entries[h])
        ws = wb.create_sheet(tre_name)
        ws.freeze_panes = "A2"

        for ci, (hdr, w) in enumerate(
            zip(["ID", "英文原文", "中文翻譯", "說明"], [8, 55, 55, 30]), 1
        ):
            c = ws.cell(row=1, column=ci, value=hdr)
            c.font = styles["hdr_font"]
            c.fill = styles["hdr_fill"]
            c.alignment = styles["wrap"]
            ws.column_dimensions[get_column_letter(ci)].width = w

        for ri, (eid, txt) in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=eid)
            ws.cell(row=ri, column=2, value=txt)
            ws.cell(row=ri, column=3, value="")
            ws.cell(row=ri, column=4, value=description if ri == 2 else "")
            for ci in range(1, 5):
                c = ws.cell(row=ri, column=ci)
                c.fill = row_fill
                c.alignment = styles["wrap"]
                c.border = styles["thin"]
            ws.row_dimensions[ri].height = 35

        total += len(rows)
        print(f"      ✅ {fname:12s} {len(rows):3d} 條")

    wb.save(out_path)
    print(f"      總計：{total} 條介面文字")


# ─── IMPORT：統一指令，讀兩個 Excel + 自動偵測兩個字型 ──────────────────────────


def cmd_import(sub_xlsx: str, ui_xlsx: str, out_dir: str, font_dir: str = None):
    """
    讀入 subtitles.xlsx + ui_text.xlsx，自動在 font_dir（預設為目前工作目錄）
    尋找 SUBFONT.TTF（字幕字型）與 UIFONT.TTF（介面字型），
    輸出：
      {out_dir}/SUBTITLES.MIX  — 包含字幕翻譯 + 兩個字型 + SBTLVERS.TRE 宣告
      {out_dir}/*.TRE          — 介面翻譯（供散落檔案覆蓋使用，若不用 patch-tlk）
    """
    try:
        import openpyxl
    except ImportError:
        print("❌ 請先安裝：pip install openpyxl")
        sys.exit(1)

    if font_dir is None:
        font_dir = os.getcwd()

    os.makedirs(out_dir, exist_ok=True)

    # ── 1. 自動偵測字型檔 ────────────────────────────────────────────────
    print(f"[1/6] 偵測字型檔（搜尋目錄：{font_dir}）...")
    sub_font_path = os.path.join(font_dir, SUB_FONT_FILENAME)
    ui_font_path = os.path.join(font_dir, UI_FONT_FILENAME)

    has_sub_font = os.path.isfile(sub_font_path)
    has_ui_font = os.path.isfile(ui_font_path)

    if has_sub_font:
        print(f"      ✅ 字幕字型：{sub_font_path}")
    else:
        print(
            f"      ⚠️  找不到字幕字型 '{SUB_FONT_FILENAME}'（請放在 {font_dir} 或用 --font-dir 指定路徑）"
        )

    if has_ui_font:
        print(f"      ✅ 介面字型：{ui_font_path}")
    else:
        print(
            f"      ⚠️  找不到介面字型 '{UI_FONT_FILENAME}'（介面將沿用字幕字型或內建點陣字型）"
        )

    if not has_sub_font and not has_ui_font:
        print(f"\n❌ 兩個字型都找不到，無法繼續。請準備：")
        print(f"      {SUB_FONT_FILENAME}  — 字幕用字型")
        print(
            f"      {UI_FONT_FILENAME}   — 介面用字型（可以跟字幕字型是同一個檔案，改名即可）"
        )
        sys.exit(1)

    # ── 2. 讀取字幕翻譯 ──────────────────────────────────────────────────
    sub_translations = {}
    total_sub_translated = 0
    if os.path.isfile(sub_xlsx):
        print(f"\n[2/6] 讀取字幕翻譯：{sub_xlsx}")
        wb = openpyxl.load_workbook(sub_xlsx, read_only=True, data_only=True)
        for prefix, typ, loc in TRE_DEFS:
            if typ in ("binary", "meta"):
                continue
            if prefix not in wb.sheetnames:
                continue
            ws = wb[prefix]
            rows_iter = iter(ws.iter_rows(values_only=True))
            next(rows_iter)
            result = []
            for row in rows_iter:
                if not row or row[0] is None:
                    continue
                if typ == "ingquo":
                    eid, en_txt, zh_txt = row[0], row[1] or "", row[2] or ""
                    eid = int(eid)
                else:
                    fs, fe = int(row[0] or 0), int(row[1] or 0)
                    en_txt = row[3] or ""
                    zh_txt = row[4] or ""
                    eid = pack_vqa_id(fs, fe)
                final = (
                    str(zh_txt).strip() if str(zh_txt).strip() else str(en_txt).strip()
                )
                if str(zh_txt).strip():
                    total_sub_translated += 1
                result.append((eid, final))
            if result:
                sub_translations[prefix] = result
                print(f"      {prefix:15s} {len(result):4d} 條")
        wb.close()
        print(f"      已翻譯：{total_sub_translated} 條")
    else:
        print(f"\n[2/6] 跳過字幕（找不到 {sub_xlsx}）")

    # ── 3. 讀取介面翻譯 ──────────────────────────────────────────────────
    ui_translations = {}
    total_ui_translated = 0
    if os.path.isfile(ui_xlsx):
        print(f"\n[3/6] 讀取介面翻譯：{ui_xlsx}")
        wb = openpyxl.load_workbook(ui_xlsx, read_only=True, data_only=True)
        for tre_name, _ in UI_TRES:
            if tre_name not in wb.sheetnames:
                continue
            ws = wb[tre_name]
            rows_iter = iter(ws.iter_rows(values_only=True))
            next(rows_iter)
            entries = []
            translated = 0
            for row in rows_iter:
                if not row or row[0] is None:
                    continue
                eid = int(row[0])
                en_txt = str(row[1] or "")
                zh_txt = str(row[2] or "").strip()
                final = zh_txt if zh_txt else en_txt
                if zh_txt:
                    translated += 1
                entries.append((eid, final))
            if entries:
                ui_translations[tre_name] = entries
                total_ui_translated += translated
                print(f"      {tre_name:12s} {translated}/{len(entries)} 條已翻譯")
        wb.close()
    else:
        print(f"\n[3/6] 跳過介面（找不到 {ui_xlsx}）")

    # ── 4. 準備 SUBTITLES.MIX 內容 ──────────────────────────────────────
    print(f"\n[4/6] 準備 SUBTITLES.MIX...")
    file_map = {}

    # 字幕 TRE
    for prefix, entries in sub_translations.items():
        loc = dict((p, l) for p, t, l in TRE_DEFS)[prefix]
        fname = tre_filename(prefix, loc)
        file_map[fname] = write_tre(entries)
        print(f"      {fname} ← 字幕翻譯")

    # SBTLVERS.TRE：宣告字幕字型 + （若有）介面字型旗標
    vers = [
        (0, "BR TW Fan Translation"),
        (1, "1"),
        (2, datetime.now().strftime("%H:%M:%S %d/%m/%Y")),
        (3, "ZH_TWN"),
        (4, "ttf" if has_sub_font else ""),
        (5, SUB_FONT_FILENAME if has_sub_font else ""),
        (6, "SIL Open Font License 1.1"),
        (7, ""),
    ]
    file_map["SBTLVERS.TRE"] = write_tre(vers)

    if has_sub_font:
        with open(sub_font_path, "rb") as f:
            file_map[SUB_FONT_FILENAME] = f.read()

    if has_ui_font:
        with open(ui_font_path, "rb") as f:
            file_map[UI_FONT_FILENAME] = f.read()
        print(
            f"      {UI_FONT_FILENAME} ← 介面專用字型（bladerunner.cpp 會優先讀取此固定檔名）"
        )

    # ── 5. 寫出 SUBTITLES.MIX ────────────────────────────────────────────
    print(f"\n[5/6] 封裝 SUBTITLES.MIX...")
    mix_path = os.path.join(out_dir, "SUBTITLES.MIX")
    with open(mix_path, "wb") as f:
        f.write(pack_mix(file_map))
    print(f"      ✅ {mix_path}（{os.path.getsize(mix_path):,} bytes）")

    # ── 6. 寫出介面 .TRE（供散落檔案覆蓋方式使用） ──────────────────────
    if ui_translations:
        print(
            f"\n[6/6] 輸出介面 .TRE 檔案（散落檔案方式，亦可改用 patch-tlk 寫回 TLK）..."
        )
        for tre_name, entries in ui_translations.items():
            out_path = os.path.join(out_dir, f"{tre_name}.TRE")
            with open(out_path, "wb") as f:
                f.write(write_tre(entries))
            print(f"      ✅ {out_path}")
    else:
        print(f"\n[6/6] 無介面翻譯可輸出")

    print(f"\n{'='*55}")
    print(f"  ✅ 完成！")
    print(f"  字幕翻譯：{total_sub_translated} 條")
    print(f"  介面翻譯：{total_ui_translated} 條")
    print(f"  輸出目錄：{out_dir}")
    print(f"  下一步：")
    print(f"    1) 把 SUBTITLES.MIX 複製到遊戲目錄")
    print(
        f"    2) 介面翻譯可選：複製 .TRE 到遊戲目錄，或執行 patch-tlk 直接寫回 TLK（建議）"
    )
    print("=" * 55)


# ─── PATCH-TLK：維持原樣，直接寫回 TLK ──────────────────────────────────────────


def cmd_patch_tlk(game_dir: str, xlsx_path: str):
    """直接把 UI 翻譯寫回 TLK 檔案（原地修改，自動備份）"""
    try:
        import openpyxl
    except ImportError:
        print("❌ 請先安裝：pip install openpyxl")
        sys.exit(1)

    UI_HASHES = {fold_hash(f"{name}.TRE"): name for name, _ in UI_TRES}

    tlk_names = ["1.TLK", "2.TLK", "3.TLK", "A.TLK"]
    tlk_map = {}

    print("[1/5] 掃描 TLK，定位各 UI TRE...")
    for tlk_name in tlk_names:
        tlk_path = os.path.join(game_dir, tlk_name)
        if not os.path.isfile(tlk_path):
            continue
        with open(tlk_path, "rb") as f:
            data = f.read()
        entries = unpack_mix(data)
        for h in entries:
            if h in UI_HASHES:
                tre_name = UI_HASHES[h]
                tlk_map[tre_name] = tlk_name
                print(f"      {tre_name:12s} → {tlk_name}")

    if not tlk_map:
        print("❌ 找不到任何 UI TRE，確認遊戲目錄正確")
        sys.exit(1)

    print(f"\n[2/5] 讀取翻譯 Excel：{xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    translated_tres = {}
    total_translated = 0

    for tre_name, _ in UI_TRES:
        if tre_name not in wb.sheetnames:
            continue
        ws = wb[tre_name]
        rows_iter = iter(ws.iter_rows(values_only=True))
        next(rows_iter)
        entries = []
        translated = 0
        for row in rows_iter:
            if not row or row[0] is None:
                continue
            eid = int(row[0])
            en_txt = str(row[1] or "")
            zh_txt = str(row[2] or "").strip()
            final = zh_txt if zh_txt else en_txt
            if zh_txt:
                translated += 1
            entries.append((eid, final))
        if entries:
            translated_tres[tre_name] = entries
            total_translated += translated
            print(f"      {tre_name:12s} {translated}/{len(entries)} 條已翻譯")
    wb.close()

    print(f"\n[3/5] 備份原始 TLK...")
    patched_tlks = set(tlk_map[n] for n in translated_tres if n in tlk_map)
    for tlk_name in patched_tlks:
        tlk_path = os.path.join(game_dir, tlk_name)
        bak_path = tlk_path + ".bak"
        if not os.path.isfile(bak_path):
            shutil.copy2(tlk_path, bak_path)
            print(f"      備份：{bak_path}")
        else:
            print(f"      已有備份：{bak_path}（跳過）")

    print(f"\n[4/5] 寫入翻譯到 TLK...")
    for tlk_name in patched_tlks:
        tlk_path = os.path.join(game_dir, tlk_name)
        with open(tlk_path, "rb") as f:
            tlk_data = f.read()

        all_entries = unpack_mix(tlk_data)

        for tre_name, entries in translated_tres.items():
            if tlk_map.get(tre_name) != tlk_name:
                continue
            h = fold_hash(f"{tre_name}.TRE")
            all_entries[h] = write_tre(entries)
            print(f"      替換 {tre_name}.TRE → {tlk_name}")

        new_tlk = repack_mix_preserve_unknown(tlk_data, all_entries)
        with open(tlk_path, "wb") as f:
            f.write(new_tlk)
        print(f"      ✅ {tlk_name} 寫入完成（{len(new_tlk):,} bytes）")

    print(f"\n{'='*55}")
    print(f"  ✅ 完成！共翻譯 {total_translated} 條")
    print(f"  直接用 ScummVM 啟動遊戲即可，不需要複製任何額外檔案")
    print(f"  若需還原：把 .bak 檔案改回原本的 .TLK 名稱即可")
    print("=" * 55)


# ─── CLI ──────────────────────────────────────────────────────────────────────


def main():
    p = argparse.ArgumentParser(description="Blade Runner 字幕 & UI 翻譯工具")
    sub = p.add_subparsers(dest="cmd", required=True)

    d = sub.add_parser("dump", help="掃描遊戲目錄，輸出 subtitles.xlsx + ui_text.xlsx")
    d.add_argument(
        "--dir", required=True, help="遊戲目錄（含 SUBTITLES.MIX、1.TLK 等）"
    )
    d.add_argument("--out", required=True, help="輸出目錄（會產生兩個 .xlsx）")

    i = sub.add_parser(
        "import", help="讀入兩個 Excel，自動偵測字型，輸出 SUBTITLES.MIX + UI 檔案"
    )
    i.add_argument("--subxlsx", required=True, help="字幕翻譯 Excel 路徑")
    i.add_argument("--uixlsx", required=True, help="介面翻譯 Excel 路徑")
    i.add_argument("--out", required=True, help="輸出目錄")
    i.add_argument(
        "--font-dir",
        default=None,
        help="字型搜尋目錄（預設為目前工作目錄），需放 SUBFONT.TTF 和/或 UIFONT.TTF",
    )

    pt = sub.add_parser(
        "patch-tlk", help="直接把 UI 翻譯寫回 TLK 檔案（最簡單，建議用這個處理介面）"
    )
    pt.add_argument("--dir", required=True, help="遊戲目錄（含 1.TLK 等）")
    pt.add_argument("--xlsx", required=True, help="介面翻譯 Excel 路徑（ui_text.xlsx）")

    args = p.parse_args()

    if args.cmd == "dump":
        if not os.path.isdir(args.dir):
            print(f"❌ 找不到目錄：{args.dir}")
            sys.exit(1)
        cmd_dump(args.dir, args.out)

    elif args.cmd == "import":
        if not os.path.isfile(args.subxlsx) and not os.path.isfile(args.uixlsx):
            print(f"❌ 找不到 {args.subxlsx} 或 {args.uixlsx}")
            sys.exit(1)
        cmd_import(args.subxlsx, args.uixlsx, args.out, args.font_dir)

    elif args.cmd == "patch-tlk":
        if not os.path.isdir(args.dir):
            print(f"❌ 找不到目錄：{args.dir}")
            sys.exit(1)
        if not os.path.isfile(args.xlsx):
            print(f"❌ 找不到：{args.xlsx}")
            sys.exit(1)
        cmd_patch_tlk(args.dir, args.xlsx)


if __name__ == "__main__":
    main()
